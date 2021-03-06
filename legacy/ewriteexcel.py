#!/usr/bin/env python

# miksi vitussa tää toimii????
# luo tiedoston 39 välilyönnillä????
# ewriteexcel - load workbook with name.xlsx
import RPi.GPIO as GPIO
import SimpleMFRC522
import openpyxl
import datetime

while True:
	inside = "in"
	outside = "out"
	nameis = ""
	direction = ""
	
	while direction != outside and direction != inside:
	
		direction = input("Going inside or outside?: ")
	
	reader = SimpleMFRC522.SimpleMFRC522()
	id, text = reader.read()
	date = datetime.date.today()
	time = datetime.datetime.now().time()
	wb = openpyxl.load_workbook(text.replace(" ", "") + '.xlsx')
	
	if text in wb.sheetnames:
		nameis = "True"
		sheet = wb[text]
	
	if nameis != "True":
		wb.create_sheet(text)
		sheet = wb [text]

#sheet = wb['Sheet1']


#while direction != outside and direction != inside:
#
#        direction = input("Going inside or outside?: ")

	if direction == inside:
	        a = sheet.max_row
		sheet.cell(row=a+1,column=1).value="Inside"
	        sheet.cell(row=a+1,column=2).value=text
	        sheet.cell(row=a+1,column=3).value=date
	        sheet.cell(row=a+1,column=4).value=time
	
	        wb.save(text + '.xlsx')
		print("Welcome, " + text)
	
		GPIO.cleanup()
	
	elif direction == outside:
		a = sheet.max_row
	        sheet.cell(row=a+1,column=5).value="Outside"
	        sheet.cell(row=a+1,column=6).value=text
	        sheet.cell(row=a+1,column=7).value=date
	        sheet.cell(row=a+1,column=8).value=time
	
	        wb.save(text + '.xlsx')
	        print("Goodbye, " + text)
	
		GPIO.cleanup()
