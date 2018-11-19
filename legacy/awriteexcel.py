#!/usr/bin/env python
#ennen openpyxl
import RPi.GPIO as GPIO
import SimpleMFRC522
import openpyxl
import datetime

try:
	reader = SimpleMFRC522.SimpleMFRC522()
	id, text = reader.read()
	date = datetime.date.today()
	time = datetime.datetime.now().time()
	wb = openpyxl.load_workbook('worktime.xlsx')
	sheet = wb['Sheet1']

	a = sheet.max_row
	sheet.cell(row=a+1,column=1).value=text
	sheet.cell(row=a+1,column=2).value=date
	sheet.cell(row=a+1,column=3).value=time

	wb.save('worktime.xlsx')
	print('Welcome ' + text)
finally:
	GPIO.cleanup()
