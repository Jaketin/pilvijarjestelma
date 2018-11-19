#!/usr/bin/env python

import RPi.GPIO as GPIO
import SimpleMFRC522
import openpyxl
import datetime

inside = "in"
outside = "out"
direction = ""
reader = SimpleMFRC522.SimpleMFRC522()
id, text = reader.read()
date = datetime.date.today()
time = datetime.datetime.now().time()
wb = openpyxl.load_workbook('worktime.xlsx')
sheet = wb['Sheet1']


while direction != outside and direction != inside:

        direction = input("Going inside or outside?: ")

if direction == inside:
        a = sheet.max_row
        sheet.cell(row=a+1,column=1).value=text
        sheet.cell(row=a+1,column=2).value=date
        sheet.cell(row=a+1,column=3).value=time

        wb.save('worktime.xlsx')
	print("Welcome, " + text)

	GPIO.cleanup()

elif direction == outside:
	a = sheet.max_row
        sheet.cell(row=a+1,column=4).value=text
        sheet.cell(row=a+1,column=5).value=date
        sheet.cell(row=a+1,column=6).value=time

        wb.save('worktime.xlsx')
        print("Goodbye, " + text)

	GPIO.cleanup()
