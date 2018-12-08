#!/usr/bin/env python
# Version 1.0.2

# import dependencies

import RPi.GPIO as GPIO
import SimpleMFRC522
import openpyxl
import datetime
import shutil

while True: 
	# keep the program running, allowing constant reading of tags without restarting
	
	# create variables
	inside = "in"
	outside = "out"
	nameis = "" # used in creating/using a worksheet into the .xlsx
	direction = "" # used in determining the direction of the event
	
	while direction != outside and direction != inside:
		
		# if direction hasn't been determined yet, ask for it
	
		direction = input("Going inside or outside?: ")
	
	reader = SimpleMFRC522.SimpleMFRC522()
	id, text = reader.read()
		# read the name on a tag
		
	date = datetime.date.today()
	time = datetime.datetime.now().time()
	wb = openpyxl.load_workbook(text.replace(" ", "") + '.xlsx')
		# in our team's RFID tag there were many spaces written at the end of the data for some reason - text.replace removes all spaces
	
	if text.replace(" ", "") in wb.sheetnames:
		# if sheet with name already exists, use it
		
		nameis = "True"
		sheet = wb[text.replace(" ", "")]
	
	if nameis != "True":
		# if sheet with name doesn't exist, create one and use it
		
		wb.create_sheet(text.replace(" ", ""))
		sheet = wb [text.replace(" ", "")]

	if direction == inside:
		# if heading to work, write all the data on columns 1-4
		
	        a = sheet.max_row
		sheet.cell(row=a+1,column=1).value="Inside"
	        sheet.cell(row=a+1,column=2).value=text
	        sheet.cell(row=a+1,column=3).value=date
	        sheet.cell(row=a+1,column=4).value=time
	
	        wb.save(text.replace(" ","") + '.xlsx')
			# save the file and copy it to NextCloud client folder
			
		shutil.copy2('/home/pi/pilvijarjestelma/' + text.replace(" ", "") + '.xlsx', '/home/pi/NextcloudFold/excel')
		print("Welcome, " + text)
	
		GPIO.cleanup()
	
	elif direction == outside:
		# if leaving work, write on columns 5-8
		
		a = sheet.max_row
	        sheet.cell(row=a+1,column=5).value="Outside"
	        sheet.cell(row=a+1,column=6).value=text
	        sheet.cell(row=a+1,column=7).value=date
	        sheet.cell(row=a+1,column=8).value=time

		sheet.cell(row=a+1,column=10).value='=SUM(OFFSET(INDIRECT(ADDRESS(ROW(); COLUMN()));0;-1;1;1)-OFFSET(INDIRECT(ADDRESS(ROW(); COLUMN()));-1;-5;1;1))'
			# calculate work time
	
	        wb.save(text.replace(" ", "") + '.xlsx')
			# save the file and copy it to NextCloud client folder
			
		shutil.copy2('/home/pi/pilvijarjestelma/' + text.replace(" ", "") + '.xlsx', '/home/pi/NextcloudFold/excel')
	        print("Goodbye, " + text)
	
		GPIO.cleanup()
